"""Validate the generated FP&A portfolio delivery."""

from __future__ import annotations

import json
import re
import sqlite3
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
QA = ROOT / "QA"
QA.mkdir(exist_ok=True)

checks: list[dict] = []


def check(name: str, passed: bool, detail: str, category: str) -> None:
    checks.append(
        {
            "category": category,
            "check": name,
            "passed": bool(passed),
            "detail": detail,
        }
    )


validation = json.loads((ROOT / "Data" / "validation_report.json").read_text())
check(
    "Financial and data controls",
    validation["all_passed"]
    and validation["passed_checks"] == validation["total_checks"] == 17,
    f'{validation["passed_checks"]}/{validation["total_checks"]} passed',
    "Analytics",
)

with sqlite3.connect(ROOT / "SQL" / "integrated_fpa_analytics.db") as conn:
    table_count = conn.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE type IN ('table','view')"
    ).fetchone()[0]
    scenario_count = conn.execute("SELECT COUNT(*) FROM scenario_summary").fetchone()[0]
check(
    "SQLite analytical database",
    table_count >= 20 and scenario_count == 4,
    f"{table_count} tables/views; {scenario_count} scenarios",
    "Analytics",
)

xlsx = ROOT / "Excel" / "Integrated_FP&A_Budgeting_Forecasting_Scenario_Planning_Model.xlsx"
workbook = load_workbook(xlsx, read_only=False, data_only=False)
formula_count = sum(
    1
    for sheet in workbook.worksheets
    for row in sheet.iter_rows()
    for cell in row
    if isinstance(cell.value, str) and cell.value.startswith("=")
)
check(
    "Excel workbook structure",
    len(workbook.sheetnames) == 31 and formula_count >= 300,
    f"{len(workbook.sheetnames)} sheets; {formula_count} formulas",
    "Excel",
)
formula_scan = json.loads((ROOT / "Excel" / "formula_error_scan.json").read_text())
check(
    "Excel formula-error scan",
    "matched 0 entries" in formula_scan["ndjson"],
    "0 formula-error matches",
    "Excel",
)

pbip_validation = json.loads(
    (ROOT / "PowerBI" / "Integrated_FPA_PBIP" / "pbip_validation.json").read_text()
)
model = json.loads(
    (
        ROOT
        / "PowerBI"
        / "Integrated_FPA_PBIP"
        / "Integrated_FPA_Planning_Analytics.SemanticModel"
        / "model.bim"
    ).read_text()
)
measure_count = sum(len(table.get("measures", [])) for table in model["model"]["tables"])
check(
    "Power BI PBIP structure",
    pbip_validation["valid"]
    and pbip_validation["pages"] == 11
    and pbip_validation["visuals"] == 89
    and measure_count >= 40,
    f'{pbip_validation["pages"]} pages; {pbip_validation["visuals"]} visuals; {measure_count} measures',
    "Power BI",
)

decks = {
    "English": ROOT
    / "Presentation"
    / "Integrated_FPA_Budgeting_Forecasting_Scenario_Planning_Professional_Deck_EN.pptx",
    "Turkish": ROOT
    / "Presentation"
    / "Entegre_FPA_Butceleme_Tahminleme_Senaryo_Planlama_Profesyonel_Sunum_TR.pptx",
}
for language, filename in decks.items():
    with zipfile.ZipFile(filename) as archive:
        names = archive.namelist()
        slide_count = len(
            [
                name
                for name in names
                if re.fullmatch(r"ppt/slides/slide\d+\.xml", name)
            ]
        )
        notes_count = len(
            [
                name
                for name in names
                if re.fullmatch(r"ppt/notesSlides/notesSlide\d+\.xml", name)
            ]
        )
    check(
        f"{language} presentation",
        slide_count == 20 and notes_count == 20,
        f"{slide_count} slides; {notes_count} speaker-note pages",
        "Presentation",
    )

pdf = ROOT / "Reports" / "Integrated_FPA_Executive_Report_12_Page_Vector_HD.pdf"
reader = PdfReader(pdf)
text_pages = sum(bool(page.extract_text().strip()) for page in reader.pages)
check(
    "Vector executive PDF",
    len(reader.pages) == 12 and text_pages == 12,
    f"{len(reader.pages)} pages; selectable text on {text_pages} pages",
    "Report",
)

core_images = [
    "executive-overview.png",
    "solution-architecture.png",
    "budget-vs-forecast.png",
    "forecast-governance.png",
    "cash-liquidity.png",
    "scenario-risk.png",
    "presentation-cover-en.png",
    "presentation-cover-tr.png",
]
sizes = {}
images_ok = True
for name in core_images:
    with Image.open(ROOT / "Images" / name) as image:
        sizes[name] = list(image.size)
        images_ok &= image.width >= 1920 and image.height >= 1080
check(
    "HD portfolio images",
    images_ok,
    f"{len(core_images)} images at 1920×1080 or higher",
    "Images",
)

linkedin = (ROOT / "Docs" / "linkedin_project_description.md").read_text()
linkedin_body = linkedin.split("## Description — under 2,000 characters\n\n", 1)[1].strip()
check(
    "LinkedIn character limit",
    len(linkedin_body) <= 2000,
    f"{len(linkedin_body)} characters",
    "Documentation",
)

readme = (ROOT / "README.md").read_text()
links = re.findall(r"\[[^\]]+\]\(([^)]+)\)", readme)
local_links = [link for link in links if not re.match(r"https?://", link)]
missing = [link for link in local_links if not (ROOT / link).exists()]
check(
    "README local links",
    not missing,
    f"{len(local_links)} links checked; {len(missing)} missing",
    "Documentation",
)

package_results = []
for filename, minimum_entries in [
    ("Integrated_FPA_Full_Project_Source_and_Deliverables.zip", 200),
    ("Integrated_FPA_Portfolio_Key_Deliverables.zip", 15),
]:
    package = ROOT / "Delivery" / filename
    if package.exists():
        with zipfile.ZipFile(package) as archive:
            corrupt = archive.testzip()
            entries = len(archive.namelist())
        package_results.append(corrupt is None and entries >= minimum_entries)
    else:
        entries = 0
        package_results.append(False)
check(
    "Delivery package integrity",
    all(package_results),
    "Full and compact ZIP packages passed CRC validation",
    "Delivery",
)

all_passed = all(item["passed"] for item in checks)
result = {
    "project": "Integrated FP&A Budgeting, Forecasting & Scenario Planning System",
    "all_passed": all_passed,
    "passed_checks": sum(item["passed"] for item in checks),
    "total_checks": len(checks),
    "checks": checks,
    "image_dimensions": sizes,
}
(QA / "delivery_validation.json").write_text(json.dumps(result, indent=2), encoding="utf-8")

lines = [
    "# Delivery Validation",
    "",
    f"**Result: {'PASS' if all_passed else 'FAIL'} — {result['passed_checks']}/{result['total_checks']} checks passed**",
    "",
    "| Category | Check | Result | Detail |",
    "|---|---|---:|---|",
]
for item in checks:
    lines.append(
        f"| {item['category']} | {item['check']} | {'PASS' if item['passed'] else 'FAIL'} | {item['detail']} |"
    )
(QA / "delivery_validation.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

print(json.dumps(result, indent=2))
if not all_passed:
    raise SystemExit(1)
